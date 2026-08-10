from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from email_validator import EmailNotValidError, validate_email
from openpyxl import load_workbook

from app.services.learner_profile_fields import (
    normalize_learner_snils,
)


HEADER_ALIASES: dict[str, set[str]] = {
    "full_name": {
        "\u0444\u0438\u043e",
        "\u0444 \u0438 \u043e",
        "\u0444\u0430\u043c\u0438\u043b\u0438\u044f \u0438\u043c\u044f \u043e\u0442\u0447\u0435\u0441\u0442\u0432\u043e",
        "full name",
        "name",
    },
    "email": {
        "email",
        "e mail",
        "\u044d\u043b\u0435\u043a\u0442\u0440\u043e\u043d\u043d\u0430\u044f \u043f\u043e\u0447\u0442\u0430",
        "\u043f\u043e\u0447\u0442\u0430",
    },
    "phone": {
        "\u0442\u0435\u043b\u0435\u0444\u043e\u043d",
        "phone",
        "\u043d\u043e\u043c\u0435\u0440 \u0442\u0435\u043b\u0435\u0444\u043e\u043d\u0430",
    },
    "program_title": {
        "\u043f\u0440\u043e\u0433\u0440\u0430\u043c\u043c\u0430",
        "\u043d\u0430\u0437\u0432\u0430\u043d\u0438\u0435 \u043f\u0440\u043e\u0433\u0440\u0430\u043c\u043c\u044b",
        "program",
    },
    "course_code": {
        "\u0443\u043a\u0434",
        "\u043a\u043e\u0434 \u043a\u0443\u0440\u0441\u0430",
        "course code",
    },
    "training_dates_text": {
        "\u0434\u0430\u0442\u044b \u043e\u0431\u0443\u0447\u0435\u043d\u0438\u044f",
        "\u043f\u0435\u0440\u0438\u043e\u0434 \u043e\u0431\u0443\u0447\u0435\u043d\u0438\u044f",
        "training dates",
    },
    "stage": {
        "\u044d\u0442\u0430\u043f",
        "stage",
    },
    "grade": {
        "\u043e\u0446\u0435\u043d\u043a\u0430",
        "\u0440\u0435\u0437\u0443\u043b\u044c\u0442\u0430\u0442",
        "grade",
    },
    "snils": {
        "\u0441\u043d\u0438\u043b\u0441",
        "\u043d\u043e\u043c\u0435\u0440 \u0441\u043d\u0438\u043b\u0441",
        "snils",
    },
    "has_snils_text": {
        "\u043d\u0430\u043b\u0438\u0447\u0438\u0435 \u0441\u043d\u0438\u043b\u0441",
    },
    "document_number": {
        "\u043d\u043e\u043c\u0435\u0440 \u0443\u0434\u043e\u0441\u0442\u043e\u0432\u0435\u0440\u0435\u043d\u0438\u044f",
        "\u043d\u043e\u043c\u0435\u0440 \u0434\u0438\u043f\u043b\u043e\u043c\u0430",
        "document number",
    },
}

REQUIRED_KEYS = {"full_name"}


@dataclass(frozen=True)
class ParsedLearnerImportRow:
    row_number: int
    status: str
    raw_data: dict[str, Any]
    normalized_data: dict[str, Any]
    validation_errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return self.status == "valid"


@dataclass(frozen=True)
class ParsedLearnerImportResult:
    filename: str
    rows: list[ParsedLearnerImportRow]

    @property
    def total_rows(self) -> int:
        return len(self.rows)

    @property
    def valid_rows(self) -> int:
        return sum(1 for row in self.rows if row.is_valid)

    @property
    def invalid_rows(self) -> int:
        return self.total_rows - self.valid_rows


def parse_learner_import_file(filename: str, content: bytes) -> ParsedLearnerImportResult:
    suffix = Path(filename).suffix.lower()

    if suffix == ".csv":
        rows = _parse_csv(content)
    elif suffix == ".xlsx":
        rows = _parse_xlsx(content)
    else:
        raise ValueError("Unsupported import file format. Use CSV or XLSX.")

    return ParsedLearnerImportResult(filename=filename, rows=rows)


def _parse_csv(content: bytes) -> list[ParsedLearnerImportRow]:
    text = _decode_csv_content(content)
    delimiter = _detect_delimiter(text)

    reader = csv.DictReader(StringIO(text), delimiter=delimiter)
    if not reader.fieldnames:
        raise ValueError("CSV file has no header row.")

    header_map = _build_header_map(reader.fieldnames)
    rows: list[ParsedLearnerImportRow] = []

    for row_number, raw_row in enumerate(reader, start=2):
        if _is_empty_mapping(raw_row):
            continue
        rows.append(_build_parsed_row(row_number=row_number, raw_row=raw_row, header_map=header_map))

    return rows


def _parse_xlsx(content: bytes) -> list[ParsedLearnerImportRow]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active

    all_rows = list(worksheet.iter_rows(values_only=True))
    header_row_index, headers = _find_xlsx_header_row(all_rows)
    header_map = _build_header_map(headers)

    rows: list[ParsedLearnerImportRow] = []

    for zero_based_index, values in enumerate(all_rows[header_row_index + 1 :], start=header_row_index + 2):
        raw_row = _xlsx_values_to_raw_row(headers, values)
        if _is_empty_mapping(raw_row):
            continue
        rows.append(_build_parsed_row(row_number=zero_based_index, raw_row=raw_row, header_map=header_map))

    return rows


def _find_xlsx_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, list[str]]:
    for index, values in enumerate(rows):
        headers = [_cell_to_text(value) for value in values]
        header_map = _build_header_map(headers)
        if REQUIRED_KEYS.issubset(header_map.keys()):
            return index, headers

    raise ValueError("XLSX file has no recognizable header row with full name column.")


def _xlsx_values_to_raw_row(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    raw_row: dict[str, Any] = {}

    for index, header in enumerate(headers):
        if not header:
            continue

        value = values[index] if index < len(values) else None
        raw_row[header] = _cell_to_text(value)

    return raw_row


def _build_parsed_row(
    *,
    row_number: int,
    raw_row: dict[str, Any],
    header_map: dict[str, str],
) -> ParsedLearnerImportRow:
    normalized = _normalize_row(raw_row, header_map)
    validation_errors = _validate_normalized_row(normalized)

    return ParsedLearnerImportRow(
        row_number=row_number,
        status="invalid" if validation_errors else "valid",
        raw_data={str(key): _cell_to_text(value) for key, value in raw_row.items() if key is not None},
        normalized_data=normalized,
        validation_errors=validation_errors,
    )


def _normalize_row(raw_row: dict[str, Any], header_map: dict[str, str]) -> dict[str, Any]:
    full_name = _clean_spaces(_get_mapped_value(raw_row, header_map, "full_name"))
    last_name, first_name, middle_name = _split_full_name(full_name)

    email = _normalize_email(_get_mapped_value(raw_row, header_map, "email"))
    phone = _normalize_phone(_get_mapped_value(raw_row, header_map, "phone"))
    snils = _normalize_snils(_get_mapped_value(raw_row, header_map, "snils"))

    return {
        "full_name": full_name,
        "last_name": last_name,
        "first_name": first_name,
        "middle_name": middle_name,
        "email": email,
        "phone": phone,
        "program_title": _clean_spaces(_get_mapped_value(raw_row, header_map, "program_title")),
        "course_code": _clean_spaces(_get_mapped_value(raw_row, header_map, "course_code")),
        "training_dates_text": _clean_spaces(_get_mapped_value(raw_row, header_map, "training_dates_text")),
        "stage": _clean_spaces(_get_mapped_value(raw_row, header_map, "stage")),
        "grade": _clean_spaces(_get_mapped_value(raw_row, header_map, "grade")),
        "snils": snils,
        "has_snils_text": _clean_spaces(_get_mapped_value(raw_row, header_map, "has_snils_text")),
        "document_number": _clean_spaces(_get_mapped_value(raw_row, header_map, "document_number")),
    }


def _validate_normalized_row(normalized: dict[str, Any]) -> list[str]:
    errors: list[str] = []

    if not normalized["full_name"]:
        errors.append("\u0424\u0418\u041e \u043e\u0431\u044f\u0437\u0430\u0442\u0435\u043b\u044c\u043d\u043e.")

    if normalized["email"] and not _is_valid_email(normalized["email"]):
        errors.append("\u041d\u0435\u043a\u043e\u0440\u0440\u0435\u043a\u0442\u043d\u044b\u0439 email.")

    if not normalized["email"]:
        errors.append("Email \u043e\u0431\u044f\u0437\u0430\u0442\u0435\u043b\u0435\u043d \u0434\u043b\u044f \u0440\u0435\u0433\u0438\u0441\u0442\u0440\u0430\u0446\u0438\u0438 \u043d\u0430 \u043f\u043e\u0440\u0442\u0430\u043b\u0435.")

    if normalized["snils"] and not re.fullmatch(r"\d{3}-\d{3}-\d{3} \d{2}", normalized["snils"]):
        errors.append("\u0421\u041d\u0418\u041b\u0421 \u0434\u043e\u043b\u0436\u0435\u043d \u0441\u043e\u0434\u0435\u0440\u0436\u0430\u0442\u044c 11 \u0446\u0438\u0444\u0440.")

    return errors


def _build_header_map(headers: list[str]) -> dict[str, str]:
    normalized_headers = {_normalize_header(header): header for header in headers if _normalize_header(header)}
    header_map: dict[str, str] = {}

    for target_key, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            normalized_alias = _normalize_header(alias)
            if normalized_alias in normalized_headers:
                header_map[target_key] = normalized_headers[normalized_alias]
                break

    return header_map


def _get_mapped_value(raw_row: dict[str, Any], header_map: dict[str, str], key: str) -> str:
    source_header = header_map.get(key)
    if not source_header:
        return ""

    return _cell_to_text(raw_row.get(source_header))


def _decode_csv_content(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue

    raise ValueError("CSV encoding is not supported. Use UTF-8 or Windows-1251.")


def _detect_delimiter(text: str) -> str:
    first_line = text.splitlines()[0] if text.splitlines() else ""
    return ";" if first_line.count(";") > first_line.count(",") else ","


def _normalize_header(value: Any) -> str:
    text = _cell_to_text(value).lower().replace("\u0451", "\u0435")
    text = re.sub(r"[^a-z\u0430-\u044f0-9]+", " ", text)
    return _clean_spaces(text)


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _clean_spaces(value: Any) -> str:
    return re.sub(r"\s+", " ", _cell_to_text(value)).strip()


def _split_full_name(full_name: str) -> tuple[str, str, str]:
    parts = full_name.split()

    if not parts:
        return "", "", ""

    last_name = parts[0]
    first_name = parts[1] if len(parts) > 1 else ""
    middle_name = " ".join(parts[2:]) if len(parts) > 2 else ""

    return last_name, first_name, middle_name


def _normalize_email(value: Any) -> str:
    return _clean_spaces(value).lower()


def _normalize_phone(value: Any) -> str:
    text = _clean_spaces(value)
    if not text:
        return ""

    digits = re.sub(r"\D+", "", text)

    if len(digits) == 11 and digits.startswith("8"):
        return "+7" + digits[1:]

    if len(digits) == 11 and digits.startswith("7"):
        return "+" + digits

    if len(digits) == 10:
        return "+7" + digits

    return text


def _normalize_snils(value: Any) -> str:
    return normalize_learner_snils(value) or ""


def _is_valid_email(value: str) -> bool:
    try:
        validate_email(value, check_deliverability=False)
    except EmailNotValidError:
        return False

    return True


def _is_empty_mapping(row: dict[Any, Any]) -> bool:
    return not any(_cell_to_text(value) for value in row.values())
